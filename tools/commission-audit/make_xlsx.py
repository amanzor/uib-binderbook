"""Write the one-sheet cross-check workbook."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

rows = json.load(open('final_rows.json'))

NAVY = '1F3864'
GREY = 'F2F2F2'
BAND = 'FAFAFA'
TIER_FILL = {'A': 'C6E0B4', 'B': 'FFE699', 'C': 'FFF2CC', 'D': 'F2F2F2'}

F = 'Arial'
wb = Workbook()
ws = wb.active
ws.title = 'June 2026 Cross-Check'

thin = Side(style='thin', color='BFBFBF')
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def put(cell, value, *, bold=False, size=10, color='000000', fill=None,
        wrap=False, align='left', fmt=None, border=False):
    c = ws[cell]
    c.value = value
    c.font = Font(name=F, size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if fill:
        c.fill = PatternFill('solid', fgColor=fill)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = box
    return c


# ── Title ─────────────────────────────────────────────────────
ws.merge_cells('A1:Q1')
put('A1', 'Jorge Castro — June 2026 Commission Cross-Check', bold=True, size=15, color=NAVY)
ws.merge_cells('A2:Q2')
put('A2', 'Every insured named on the June 2026 commission PDFs in Admin ▸ Commission Statements '
          'who also appears on Jorge Castro\'s June 2026 commission. Work top-down: Rank A is confirmed, '
          'Rank D is almost certainly noise.', size=9.5, color='555555')
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 14

# ── What was checked ──────────────────────────────────────────
put('A4', 'WHAT WAS CHECKED', bold=True, size=10, color=NAVY)

scope = [
    ('June 2026 PDFs uploaded', '5 files, 4 unique — "Amwins Comm.PDF" was uploaded twice (byte-identical)'),
    ('  1. Amwins Specialty Auto', '01_Amwins_Comm.PDF — 9 rows, 7 insureds, $72.34 — ties to statement total'),
    ('  2. Progressive', '02_DetailedStatement20260717.pdf — 188 rows, 138 insureds, $14,289.02 — ties to statement total'),
    ('  3. Kemper Auto (Infinity)', '03_07b3dbfb-…-8a1869835bb5.pdf — 27 rows, 20 insureds, $656.60 — ties to statement total'),
    ('  4. GEICO', '04_Commission Statement (35).pdf — 38 rows, 19 insureds, $2,267.79 — ties to statement total'),
    ('Jorge Castro\'s June 2026 commission',
     '56 statement rows / 35 insureds — United Auto (52 rows) + Ocean Harbor (4 rows), $332.59 net; '
     'plus 1 binder-book entry he wrote in June (Marlene Castillo)'),
    ('Producers named on the 4 PDFs',
     'Alberto Manzor, Amanda Montano, Dahil R. Duenas, Unassigned — Jorge Castro is not a producer on any of them'),
]
r = 5
for label, detail in scope:
    put(f'A{r}', label, bold=not label.startswith('  '), size=9)
    ws.merge_cells(f'C{r}:Q{r}')
    put(f'C{r}', detail, size=9, color='333333')
    r += 1

# ── Legend ────────────────────────────────────────────────────
r += 1
put(f'A{r}', 'HOW TO READ THE RANK', bold=True, size=10, color=NAVY)
r += 1
legend = [
    ('A', 'Confirmed — the same policy number appears on both sides. No judgement needed.'),
    ('B', 'Likely the same person — surname and first name (or initial) both agree.'),
    ('C', 'Needs your eyes — surname agrees but the first name differs or is not printed.'),
    ('D', 'Common surname only and the first names differ — most likely a different person.'),
]
legend_start = r
for tier, text in legend:
    put(f'A{r}', tier, bold=True, size=9, align='center', fill=TIER_FILL[tier], border=True)
    ws.merge_cells(f'C{r}:Q{r}')
    put(f'C{r}', text, size=9, color='333333')
    r += 1

# ── Counts (formulas over the table below) ────────────────────
r += 1
put(f'A{r}', 'CANDIDATES', bold=True, size=10, color=NAVY)
count_row = r + 1                      # one count row per rank, then a gap
r = count_row + len(legend) + 1

# ── Table ─────────────────────────────────────────────────────
headers = [
    ('#', 5), ('OK?', 6), ('Rank', 6), ('What it looks like', 42),
    ('Name on the June PDF', 26), ('Carrier / statement', 20), ('Policy # on PDF', 18),
    ('Rows', 6), ('Commission on PDF', 15), ('Producer on PDF', 22),
    ("Name on Jorge Castro's June commission", 30), ('Where it sits', 26),
    ('Policy # on his commission', 20), ('Rows', 6), ('His commission', 14),
    ('Matched on', 16),
]
hdr = r
for i, (h, w) in enumerate(headers, start=1):
    put(f'{get_column_letter(i)}{hdr}', h, bold=True, size=9, color='FFFFFF',
        fill=NAVY, wrap=True, align='center', border=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hdr].height = 30

MONEY = '$#,##0.00;($#,##0.00);-'
first = hdr + 1
for n, d in enumerate(rows):
    rr = first + n
    band = BAND if n % 2 else None
    put(f'A{rr}', n + 1, size=9, align='center', fill=band, border=True)
    put(f'B{rr}', '', size=9, align='center', fill='FFFFFF', border=True)
    put(f'C{rr}', d['tier'], bold=True, size=9, align='center',
        fill=TIER_FILL[d['tier']], border=True)
    put(f'D{rr}', d['note'], size=9, wrap=True, fill=band, border=True)
    put(f'E{rr}', d['pdf_name'], bold=d['tier'] in 'AB', size=9, fill=band, border=True)
    put(f'F{rr}', d['carrier'], size=9, fill=band, border=True)
    put(f'G{rr}', d['pdf_pol'], size=9, fill=band, border=True)
    put(f'H{rr}', d['pdf_rows'], size=9, align='center', fill=band, border=True)
    put(f'I{rr}', d['pdf_amt'], size=9, align='right', fmt=MONEY, fill=band, border=True)
    put(f'J{rr}', d['producer'], size=9, fill=band, border=True)
    put(f'K{rr}', d['j_name'], bold=d['tier'] in 'AB', size=9, fill=band, border=True)
    put(f'L{rr}', d['j_src'], size=9, fill=band, border=True)
    put(f'M{rr}', d['j_pol'], size=9, fill=band, border=True)
    put(f'N{rr}', d['j_rows'], size=9, align='center', fill=band, border=True)
    put(f'O{rr}', d['j_amt'], size=9, align='right', fmt=MONEY, fill=band, border=True)
    put(f'P{rr}', d['matched_on'], size=9, fill=band, border=True)

last = first + len(rows) - 1

# Footnote. The amount columns are deliberately not totalled: one PDF row can
# pair with more than one name on Jorge's commission, so a sum double-counts.
tr = last + 2
ws.merge_cells(f'A{tr}:P{tr}')
put(f'A{tr}', f'{len(rows)} candidate pairs. Amounts are shown per pair — a single PDF line can appear '
              f'against more than one name, so these columns are not totalled. Commission figures for all four '
              f'PDFs were reconciled to the totals printed on the statements themselves.',
    size=9, color='555555')

# Rank counts, as formulas over the table.
for i, (tier, _) in enumerate(legend):
    cr = count_row + i
    put(f'C{cr}', f'=COUNTIF($C${first}:$C${last},"{tier}")', bold=True, size=9,
        align='center', fill=TIER_FILL[tier], border=True)
    ws.merge_cells(f'D{cr}:F{cr}')
    put(f'D{cr}', f'Rank {tier}', size=9)

# ── Print setup: one page wide ────────────────────────────────
ws.freeze_panes = f'A{first}'
ws.auto_filter.ref = f'A{hdr}:P{last}'
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_title_rows = f'{hdr}:{hdr}'

out = 'Jorge_Castro_June_2026_Commission_CrossCheck.xlsx'
wb.save(out)
print('wrote', out, 'rows', len(rows), 'table', first, '-', last)
