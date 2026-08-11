import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

B=json.load(open('binder.json')); C=json.load(open('carriers.json'))
BOOKM=[t for t in json.load(open('bookmatch.json'))['matched'] if not t['in_june']]
G=json.load(open('geico.json')); D2=json.load(open('carriers2.json'))

FONT='Arial'; NAVY='1F3864'
HDR=PatternFill('solid',fgColor='1F3864'); SUB=PatternFill('solid',fgColor='D9E1F2')
SECT=PatternFill('solid',fgColor='8EA9DB'); WARN=PatternFill('solid',fgColor='FFF2CC')
TOT=PatternFill('solid',fgColor='FCE4D6'); GREEN=PatternFill('solid',fgColor='E2EFDA')
thin=Side(style='thin',color='BFBFBF'); box=Border(left=thin,right=thin,top=thin,bottom=thin)
BLUE=Font(name=FONT,size=10,color='0000FF')
MONEY='$#,##0.00;($#,##0.00);-'; PCT='0.00%'
PROMO='Promotional Incentive'

def norm(p):
    p=re.sub(r'[^A-Z0-9]','',str(p).upper())
    m=re.match(r'^([A-Z]*)(\d*)$',p)
    if m and m.group(2): return m.group(1)+m.group(2).lstrip('0')
    return p

STMT={'Progressive':'Progressive Detailed Statement 06/2026',
      'Infinity':'Kemper Auto Monthly Producer Stmt 06/2026',
      'AmWins':'Amwins Specialty Auto Commission Stmt 06/2026',
      'United Auto':'United Insurance Group Direct Billing Stmt 06/2026',
      'Ocean Harbor':'Pearl Holding Group Commission Stmt 06/2026',
      'National General':'National General Commission Stmt 06/2026'}

wb=Workbook()

# ================= Transaction Detail =================
d=wb.active; d.title='Transaction Detail'
dcols=['Match Key','Carrier','Commission Statement','Statement Policy #','Insured Name (per statement)',
       'Transaction Type','Transaction Date','Written / Gross Premium','Commissionable Basis',
       'Carrier Comm Rate','Commission per Statement','Check: Basis x Rate','Diff',
       'Premium Counted in Summary','Commission Counted in Summary',
       'In Doral Binder Book?','Binder Customer Name']
d.append(dcols)

binder_keys={}
for b in B:
    b['_key']=f"{b['co']}|{norm(b['policy'])}"
    binder_keys.setdefault(b['_key'],b['name'])

drows=[]
def add(carrier,policy,name,ttype,tdate,written,basis,rate,comm):
    key=f"{carrier}|{norm(policy)}"
    drows.append([key,carrier,STMT[carrier],str(policy),name,ttype,tdate,written,basis,rate,comm,
                  None,None,None,None,'YES' if key in binder_keys else 'No',binder_keys.get(key,'')])

for p in C['prog']:
    add('Progressive',p['policy'],p['name'],p['tran'],p['tdate'],p['prem'],p['prem'],p['rate'],p['comm'])
for k in C['kemper']:
    add('Infinity',k['acct'],k['name'],k['tran'],k['date'],k['prem'],k['prem'],k['rate'],k['comm'])
for a in C['amwins']:
    add('AmWins',a['policy'],a['name'],'Commission',a['eff'],a['prem'],a['net'],a['rate'],a['comm'])
TT={'COMM AS COLLECTED':'Comm as Collected','NEW BUSINESS':'New Business','ENDORSEMENT':'Endorsement',
    'CANCEL (PRORATE)':'Cancel (Prorate)','PROMOTIONAL INCENTIVE':PROMO,'REINSTATEMENT':'Reinstatement'}
for u in D2['united']:
    add('United Auto',u['policy'],u['name'],TT.get(u['desc'],u['desc'].title()),u['date'],
        u['prem'],u['prem'],u['rate'],u['comm'])
for o in D2['ocean']:
    act={'REN':'Renewal','CANC':'Cancellation','NEW':'New Business','END':'Endorsement',
         'UCC':'Uncollected Premium','REI':'Reinstatement','UCR':'Uncollected Reimb.'}.get(o['action'],o['action'])
    add('Ocean Harbor',o['policy'],o['name'],act,o['eff'],o['prem']+o['fee'],o['prem'],o['rate'],o['comm'])
for g in D2['natgen']:
    add('National General',g['policy'],g['insured'],g['tran'],g['eff'],g['prem'],g['prem'],g['rate'],g['comm'])

GS='GEICO Commission Statement 06/2026'
for g in G:
    key='GEICO|'+norm(g['policy'])
    drows.append([key,'GEICO',GS,g['policy'],g['name'],g['section']+' Commission',g['tdate'],
                  g['prem'],g['prem'],g['rate'],g['comm'],None,None,None,None,'No',''])

order={'Progressive':0,'Infinity':1,'United Auto':2,'Ocean Harbor':3,'National General':4,'AmWins':5,'GEICO':6}
drows.sort(key=lambda r:(order[r[1]], r[15]!='YES', r[3], r[6]))
for r in drows: d.append(r)
n=len(drows)
for i in range(2,n+2):
    d.cell(i,12).value=f'=ROUND(I{i}*J{i},2)'
    d.cell(i,13).value=f'=K{i}-L{i}'
    d.cell(i,14).value=f'=IF(F{i}="{PROMO}",0,I{i})'
    d.cell(i,15).value=f'=IF(F{i}="{PROMO}",0,K{i})'
tr=n+3
d.cell(tr,1).value='STATEMENT TOTAL (all transactions, all six carriers + GEICO)'
for col in (8,11,12,13,14,15):
    L=get_column_letter(col); d.cell(tr,col).value=f'=SUM({L}2:{L}{n+1})'
d.cell(tr+1,1).value='Of which: Doral binder-book clients'
for col in (8,11,12,13,14,15):
    L=get_column_letter(col)
    d.cell(tr+1,col).value=f'=SUMIF($P$2:$P${n+1},"YES",{L}2:{L}{n+1})'

# ================= Book Activity by Carrier =================
ADJ=json.load(open('adjustments.json'))
ba=wb.create_sheet('Book Activity by Carrier')
ba.cell(1,1).value='Other Doral Book Activity on the June 2026 Carrier Statements'
ba.cell(2,1).value=('Transactions belonging to the wider Doral office book of business that appear on the June carrier '
                    'statements but are NOT on the June binder sheet - endorsements, cancellations, adjustments and '
                    'as-collected commission on policies written in earlier months.')
bcols=['Customer Name (per Doral book)','Policy Number','Insured Name (per statement)','Transaction Type',
       'Transaction Date','Premium / Basis','Premium Counted','Carrier Comm Rate','Commission per Statement',
       'Commission Counted','Book Month','Company (per book)']
NB=len(bcols)
BA_HDR=4
for j,c in enumerate(bcols,1): ba.cell(BA_HDR,j).value=c
carrier_order=['Progressive','Infinity','United Auto','Ocean Harbor','National General','AmWins','GEICO']
br=BA_HDR+1; ba_sub={}
for co in carrier_order:
    rows=sorted([t for t in BOOKM if t['carrier']==co],key=lambda x:(x['policy'],str(x['tdate'])))
    if not rows: continue
    ba.cell(br,1).value=f'{co.upper()}  -  {len(rows)} transaction{"s" if len(rows)!=1 else ""}'
    ba.cell(br,1).font=Font(name=FONT,bold=True,color='FFFFFF')
    for j in range(1,NB+1): ba.cell(br,j).fill=SECT
    st=br+1; br+=1
    for t in rows:
        ba.cell(br,1).value=t['book_name']; ba.cell(br,2).value=t['policy']
        ba.cell(br,3).value=t['name'];      ba.cell(br,4).value=t['ttype']
        ba.cell(br,5).value=str(t['tdate'])
        ba.cell(br,6).value=t['basis']
        ba.cell(br,7).value=f'=IF(D{br}="{PROMO}",0,F{br})'
        ba.cell(br,8).value=t['rate']
        ba.cell(br,9).value=t['comm']
        ba.cell(br,10).value=f'=IF(D{br}="{PROMO}",0,I{br})'
        ba.cell(br,11).value=t['book_sheet'].strip(); ba.cell(br,12).value=t['book_co']
        br+=1
    ba.cell(br,1).value=f'{co} subtotal'
    for col in (6,7,9,10): ba.cell(br,col).value=f'=SUM({get_column_letter(col)}{st}:{get_column_letter(col)}{br-1})'
    for j in range(1,NB+1): ba.cell(br,j).fill=SUB; ba.cell(br,j).font=Font(name=FONT,bold=True)
    ba_sub[co]=br; br+=2
BA_TOT=br
ba.cell(br,1).value='TOTAL - OTHER DORAL BOOK ACTIVITY'
for col in (6,7,9,10):
    L=get_column_letter(col)
    ba.cell(br,col).value='='+'+'.join(f'{L}{v}' for v in ba_sub.values())
for j in range(1,NB+1): ba.cell(br,j).fill=TOT; ba.cell(br,j).font=Font(name=FONT,bold=True,size=11)
BA_LASTTX=br
br+=3

# ----- statement-level adjustments & fees -----
ADJ_START=br
ba.cell(br,1).value='STATEMENT-LEVEL ADJUSTMENTS & FEES'
ba.cell(br,1).font=Font(name=FONT,size=12,bold=True,color=NAVY); br+=1
ba.cell(br,1).value=('Chargebacks and fees the carriers apply outside the policy transaction listing. These carry no '
                     'policy number, so they are matched by name where the carrier gives one. Only the rows marked '
                     'YES are counted as Doral.')
ba.cell(br,1).font=Font(name=FONT,size=10,italic=True,color='404040')
ba.merge_cells(start_row=br,start_column=1,end_row=br,end_column=NB)
ba.cell(br,1).alignment=Alignment(vertical='center',wrap_text=True); ba.row_dimensions[br].height=28; br+=1
acols=['Name on the adjustment','Reference','Carrier','Adjustment Type','Date','Amount','','','','Doral?','Book Month','How it was matched']
ADJ_HDR=br
for j,c in enumerate(acols,1): ba.cell(br,j).value=c
for j in range(1,NB+1):
    c=ba.cell(br,j); c.fill=HDR; c.font=Font(name=FONT,size=10,bold=True,color='FFFFFF')
    c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center'); c.border=box
ba.row_dimensions[br].height=30; br+=1
adj_start=br
for a in sorted(ADJ,key=lambda x:(carrier_order.index(x['carrier']),not x['doral'],x['name'])):
    ba.cell(br,1).value=a['book_name'] or a['name']
    ba.cell(br,2).value=a['ref']; ba.cell(br,3).value=a['carrier']
    ba.cell(br,4).value=a['ttype']; ba.cell(br,5).value=a['date']
    ba.cell(br,6).value=a['amt']
    ba.cell(br,10).value='YES' if a['doral'] else 'No'
    ba.cell(br,11).value=a['book_sheet']; ba.cell(br,12).value=a['basis']
    if not a['doral']:
        for j in range(1,NB+1): ba.cell(br,j).fill=WARN
    br+=1
adj_end=br-1
ba.cell(br,1).value='All adjustments on the statements'
ba.cell(br,6).value=f'=SUM(F{adj_start}:F{adj_end})'
for j in range(1,NB+1): ba.cell(br,j).fill=SUB; ba.cell(br,j).font=Font(name=FONT,bold=True)
br+=1
ADJ_DORAL=br
ba.cell(br,1).value='OF WHICH DORAL - counted in the totals'
ba.cell(br,6).value=f'=SUMIF(J{adj_start}:J{adj_end},"YES",F{adj_start}:F{adj_end})'
for j in range(1,NB+1): ba.cell(br,j).fill=TOT; ba.cell(br,j).font=Font(name=FONT,bold=True,size=11)
ADJ_LAST=br
_S="'Book Activity by Carrier'!"
def adj_carrier(co):
    return (f'=SUMIFS({_S}$F${adj_start}:$F${adj_end},{_S}$C${adj_start}:$C${adj_end},"{co}",'
            f'{_S}$J${adj_start}:$J${adj_end},"YES")')

# ================= Commission Summary =================
s=wb.create_sheet('Commission Summary',0)
title=['Doral Office Binder Book - June 2026 Commission Statement',
       'Universal Insurance Brokers  |  Producer: Jorge  |  Statement period 06/01/2026 - 06/30/2026',
       'Commission taken from each carrier commission statement at the rate that carrier shows, matched by policy number. United Auto is booked at 10% as collected.']
for i,t in enumerate(title,1): s.cell(i,1).value=t
scols=['Producer','Customer Name','Binder Status','New / Renewal','Line of Business','Carrier',
       'Policy Number','Eff. Date','Term (mo)','Binder Premium','Carrier Statement Used','# Txns',
       'Carrier Comm Rate','Commissionable Premium','Commission Earned','Effective Rate','Notes']
HDR_ROW=5
for j,c in enumerate(scols,1): s.cell(HDR_ROW,j).value=c

def hits_for(b): return [r for r in drows if r[0]==b['_key']]
def pct(x): return f'{x*100:g}%'

r=HDR_ROW+1; sections=[]
for ptype,label in (('New','NEW BUSINESS'),('Renewal','RENEWALS')):
    s.cell(r,1).value=label; s.cell(r,1).font=Font(name=FONT,bold=True,color='FFFFFF')
    for j in range(1,len(scols)+1): s.cell(r,j).fill=SECT
    start=r+1; r+=1
    for b in [x for x in B if x['ptype']==ptype]:
        h=hits_for(b); co=b['co']; note=''
        base=sorted({x[9] for x in h if x[5]!=PROMO}); inc=sorted({x[9] for x in h if x[5]==PROMO})
        if not h:
            if b['prem']==0:
                note=f'12-month policy effective {b["eff"]}; no June 2026 activity on the {co} statement, so $0 is correct for this month.'
            else:
                note=f'Policy not found on the June 2026 {co} statement - commission likely paid in a different statement period.'
            rate=None
        elif len(base)==1: rate=base[0]
        else: rate=' / '.join(pct(x) for x in sorted(base,reverse=True))
        if len([x for x in h if x[5]!=PROMO])>1:
            note='; '.join(f"{x[5]} {x[7]:,.2f} @ {pct(x[9])}" for x in h if x[5]!=PROMO)
        if inc:
            p=sum(x[10] for x in h if x[5]==PROMO)
            note=(f'Booked at 10% as collected. United also paid a separate 3% promotional incentive of '
                  f'${p:,.2f} on this policy - it is shown on the Transaction Detail tab but is deliberately '
                  f'EXCLUDED from this total. | ')+note
        if h and co=='United Auto' and all(x[5]=='Comm as Collected' for x in h if x[5]!=PROMO):
            note=('United bills this policy direct and pays AS COLLECTED, so June earns commission only on the '
                  'premium United collected in June - more follows in later months. | ')+note
        s.cell(r,1).value='Jorge'; s.cell(r,2).value=b['name']; s.cell(r,3).value=b['status']
        s.cell(r,4).value='New Business' if ptype=='New' else 'Renewal'
        s.cell(r,5).value=b['lob']; s.cell(r,6).value=co; s.cell(r,7).value=b['policy']
        s.cell(r,8).value=b['eff']; s.cell(r,9).value=int(b['term']) if b['term'] else None
        s.cell(r,10).value=b['prem'] if b['prem'] else None
        s.cell(r,11).value=STMT[co]
        s.cell(r,12).value=f"=COUNTIF('Transaction Detail'!$A$2:$A${n+1},$R{r})"
        s.cell(r,13).value=rate
        s.cell(r,14).value=f"=SUMIF('Transaction Detail'!$A$2:$A${n+1},$R{r},'Transaction Detail'!$N$2:$N${n+1})"
        s.cell(r,15).value=f"=SUMIF('Transaction Detail'!$A$2:$A${n+1},$R{r},'Transaction Detail'!$O$2:$O${n+1})"
        s.cell(r,16).value=f'=IF(N{r}=0,"",O{r}/N{r})'
        s.cell(r,17).value=note; s.cell(r,18).value=b['_key']
        r+=1
    end=r-1
    s.cell(r,2).value=f'{label} SUBTOTAL'
    for col in (10,14,15): s.cell(r,col).value=f'=SUM({get_column_letter(col)}{start}:{get_column_letter(col)}{end})'
    s.cell(r,16).value=f'=IF(N{r}=0,"",O{r}/N{r})'
    sections.append((label,start,end,r))
    for j in range(1,len(scols)+1): s.cell(r,j).fill=SUB; s.cell(r,j).font=Font(name=FONT,bold=True)
    r+=2
jr=r
s.cell(jr,2).value='JUNE BINDER SHEET TOTAL'
for col in (10,14,15):
    L=get_column_letter(col); s.cell(jr,col).value=f'={L}{sections[0][3]}+{L}{sections[1][3]}'
s.cell(jr,16).value=f'=IF(N{jr}=0,"",O{jr}/N{jr})'
for j in range(1,len(scols)+1): s.cell(jr,j).fill=SUB; s.cell(jr,j).font=Font(name=FONT,bold=True,size=11)
r=jr+2

BAR=f"'Book Activity by Carrier'!$B${BA_HDR+1}:$B${BA_LASTTX}"
BAP=f"'Book Activity by Carrier'!$G${BA_HDR+1}:$G${BA_LASTTX}"
BAC=f"'Book Activity by Carrier'!$J${BA_HDR+1}:$J${BA_LASTTX}"
s.cell(r,1).value='OTHER DORAL BOOK ACTIVITY  -  policies from the wider book with June statement activity'
s.cell(r,1).font=Font(name=FONT,bold=True,color='FFFFFF')
for j in range(1,len(scols)+1): s.cell(r,j).fill=SECT
bstart=r+1; r+=1
groups={}
for t in BOOKM: groups.setdefault((t['carrier'],t['policy']),[]).append(t)
for (co,pol),ts in sorted(groups.items(),key=lambda kv:(carrier_order.index(kv[0][0]),kv[0][1])):
    base=sorted({x['rate'] for x in ts if x['ttype']!=PROMO})
    s.cell(r,1).value='Jorge'
    s.cell(r,2).value=ts[0]['book_name']
    s.cell(r,3).value=ts[0]['book_sheet'].strip()
    s.cell(r,4).value='Book activity'
    s.cell(r,5).value=ts[0]['book_co'] or ''
    s.cell(r,6).value=co
    s.cell(r,7).value=pol
    s.cell(r,11).value=STMT.get(co,GS)
    s.cell(r,12).value=f'=COUNTIF({BAR},$G{r})'
    s.cell(r,13).value=(base[0] if len(base)==1 else ' / '.join(pct(x) for x in sorted(base,reverse=True)))
    s.cell(r,14).value=f'=SUMIF({BAR},$G{r},{BAP})'
    s.cell(r,15).value=f'=SUMIF({BAR},$G{r},{BAC})'
    s.cell(r,16).value=f'=IF(N{r}=0,"",O{r}/N{r})'
    s.cell(r,17).value='; '.join(f"{x['ttype']} {x['basis']:,.2f} @ {pct(x['rate'])}" for x in ts if x['ttype']!=PROMO)
    r+=1
bend=r-1
s.cell(r,2).value='OTHER BOOK ACTIVITY SUBTOTAL'
for col in (14,15): s.cell(r,col).value=f'=SUM({get_column_letter(col)}{bstart}:{get_column_letter(col)}{bend})'
s.cell(r,16).value=f'=IF(N{r}=0,"",O{r}/N{r})'
for j in range(1,len(scols)+1): s.cell(r,j).fill=SUB; s.cell(r,j).font=Font(name=FONT,bold=True)
BOOK_SUB=r; r+=2

s.cell(r,1).value='STATEMENT ADJUSTMENTS & FEES  -  Doral-attributable only'
s.cell(r,1).font=Font(name=FONT,bold=True,color='FFFFFF')
for j in range(1,len(scols)+1): s.cell(r,j).fill=SECT
astart=r+1; r+=1
for co in carrier_order:
    rows=[a for a in ADJ if a['carrier']==co]
    if not rows: continue
    dor=[a for a in rows if a['doral']]
    s.cell(r,1).value='Jorge'
    s.cell(r,2).value=f'{co} chargebacks and fees'
    s.cell(r,4).value='Adjustment'
    s.cell(r,6).value=co
    s.cell(r,11).value=STMT.get(co,GS)
    s.cell(r,12).value=len(dor)
    s.cell(r,15).value=adj_carrier(co)
    types=sorted({a['ttype'] for a in rows})
    s.cell(r,17).value=(f"{len(dor)} of {len(rows)} rows matched to the Doral book ({', '.join(types)}). "
                        f"All {len(rows)} are listed on the Book Activity tab; the rest belong to other offices "
                        f"or cannot be attributed.")
    r+=1
aend=r-1
s.cell(r,2).value='ADJUSTMENTS SUBTOTAL'
s.cell(r,15).value=f'=SUM(O{astart}:O{aend})'
for j in range(1,len(scols)+1): s.cell(r,j).fill=SUB; s.cell(r,j).font=Font(name=FONT,bold=True)
ADJ_SUB=r; r+=2

gr=r
s.cell(gr,2).value='TOTAL DORAL JUNE COMMISSION'
s.cell(gr,10).value=f'=J{jr}'
s.cell(gr,14).value=f'=N{jr}+N{BOOK_SUB}'
s.cell(gr,15).value=f'=O{jr}+O{BOOK_SUB}+O{ADJ_SUB}'
s.cell(gr,16).value=f'=IF(N{gr}=0,"",O{gr}/N{gr})'
s.cell(gr,17).value=(f'June binder sheet + other Doral book activity + Doral-attributable statement adjustments.')
for j in range(1,len(scols)+1): s.cell(gr,j).fill=TOT; s.cell(gr,j).font=Font(name=FONT,bold=True,size=12)
SUM_LAST=gr
JUNE_LAST=jr-1   # recap must look only at the June binder rows, not the added sections

# ================= Carrier Recap =================
rc=wb.create_sheet('Carrier Recap')
rc.cell(1,1).value='Recap by Carrier - Doral Binder Book vs. Full Carrier Statement'
rcols=['Carrier','Commission Statement','Rate the Carrier Shows','Binder Policies','Found on Statement',
       'Commissionable Premium (binder clients)','Commission Earned (binder clients)',
       'Total Commission on Statement (all agency business)','Binder Share of Statement',
       'Other Book Activity - Txns','Other Book Activity - Commission',
       'Statement Adjustments (Doral)','Total Doral Commission']
for j,c in enumerate(rcols,1): rc.cell(3,j).value=c
carriers=['Progressive','Infinity','United Auto','Ocean Harbor','National General','AmWins']
rateshow={'Progressive':'8% - 14%, varies by policy','Infinity':'10% flat',
          'United Auto':'10% as collected (3% incentive excluded)','Ocean Harbor':'11% (13% on some policies)',
          'National General':'10% flat','AmWins':'10% of net cash'}
stmtshow={c:STMT[c] for c in carriers}
rr=4
for co in carriers:
    rc.cell(rr,1).value=co; rc.cell(rr,2).value=stmtshow[co]; rc.cell(rr,3).value=rateshow[co]
    rc.cell(rr,4).value=len([b for b in B if b['co']==co])
    rc.cell(rr,5).value=(f"=SUMPRODUCT(('Commission Summary'!$F${HDR_ROW+1}:$F${JUNE_LAST}=$A{rr})"
                         f"*('Commission Summary'!$L${HDR_ROW+1}:$L${JUNE_LAST}>0))")
    rc.cell(rr,6).value=f"=SUMIF('Commission Summary'!$F${HDR_ROW+1}:$F${JUNE_LAST},$A{rr},'Commission Summary'!$N${HDR_ROW+1}:$N${JUNE_LAST})"
    rc.cell(rr,7).value=f"=SUMIF('Commission Summary'!$F${HDR_ROW+1}:$F${JUNE_LAST},$A{rr},'Commission Summary'!$O${HDR_ROW+1}:$O${JUNE_LAST})"
    rc.cell(rr,8).value=f"=SUMIF('Transaction Detail'!$B$2:$B${n+1},$A{rr},'Transaction Detail'!$K$2:$K${n+1})"
    rc.cell(rr,9).value=f'=IF(H{rr}=0,"",G{rr}/H{rr})'
    rc.cell(rr,10).value=len([t for t in BOOKM if t['carrier']==co])
    rc.cell(rr,11).value=(f"='Book Activity by Carrier'!J{ba_sub[co]}" if co in ba_sub else 0)
    rc.cell(rr,12).value=adj_carrier(co)
    rc.cell(rr,13).value=f'=G{rr}+K{rr}+L{rr}'

    rr+=1
rc.cell(4+carriers.index('United Auto'),8).comment=Comment(
  'The United statement NETS to $444.33 for the whole agency because large cancellations on non-Doral '
  'policies (Charlie Hoz, Gabriela Liriano, Mirloude Petit Frere) charge back more than $1,300. This column '
  'also includes United\'s 3% promotional incentive, which the Doral totals deliberately exclude. Both are '
  'why the Doral share of this statement reads over 100% - that is correct, not an error.','Analysis')
rc.cell(rr,1).value='GEICO'; rc.cell(rr,2).value=GS; rc.cell(rr,3).value='10% / 12% / 15%'
rc.cell(rr,4).value=0; rc.cell(rr,5).value=0; rc.cell(rr,6).value=0; rc.cell(rr,7).value=0
rc.cell(rr,8).value=f"=SUMIF('Transaction Detail'!$B$2:$B${n+1},$A{rr},'Transaction Detail'!$K$2:$K${n+1})"
rc.cell(rr,10).value=len([t for t in BOOKM if t['carrier']=='GEICO'])
rc.cell(rr,11).value=(f"='Book Activity by Carrier'!J{ba_sub['GEICO']}" if 'GEICO' in ba_sub else 0)
rc.cell(rr,12).value=adj_carrier('GEICO')
rc.cell(rr,13).value=f'=G{rr}+K{rr}+L{rr}'
rc.cell(rr,1).comment=Comment('No client on the June binder sheet appears on the GEICO statement, but the wider '
  'Doral book does: Iris Lopez Sanchez (policy 6219586788) has two renewal-year lines on it. The rest of the '
  'statement covers writing agents Alberto Manzor Jr and Amanda Montano.','Analysis')
grr=rr+1
rc.cell(grr,1).value='TOTAL'
for col in (4,5,6,7,8,10,11,12,13):
    L=get_column_letter(col); rc.cell(grr,col).value=f'=SUM({L}4:{L}{rr})'
rc.cell(grr,9).value=f'=IF(H{grr}=0,"",G{grr}/H{grr})'
for j in range(1,len(rcols)+1): rc.cell(grr,j).fill=TOT; rc.cell(grr,j).font=Font(name=FONT,bold=True)
rc.cell(3,10).comment=Comment('Doral book transactions found on the June statements that are NOT on the June '
  'binder sheet - listed in full on the Book Activity by Carrier tab.','Analysis')

# ================= Notes =================
nt=wb.create_sheet('Notes & Sources')
notes=[
 ('Source files',''),
 ('June binder sheet','Doral_Office_Binder_Book_June_Sheet.pdf - 63 policies (18 new business, 45 renewals), producer Jorge. This drives the Commission Summary tab.'),
 ('Full book of business','Doral_Office_Binder_Book.xlsx - the whole Doral office book, 54 monthly sheets from Dec 2022 to Aug 2026, 2,706 policy rows and 1,383 distinct policy numbers. This drives the Book Activity by Carrier tab.'),
 ('Progressive','DetailedStatement20260811_Progressive.xlsx - 188 transactions, agent 24258, month end 202606. Ties to the Summary tab of that file: $14,289.02.'),
 ('Infinity','Kemper.pdf - Kemper Auto Monthly Producer Statement, agent 5517897. 27 transactions, ties to the FL Commission Total: $656.60.'),
 ('United Auto','United_100208_June_statement.PDF - United Insurance Group, agent 001-1D-100208, premium period 06/01/26-06/30/26. 98 transactions netting $444.33, which ties to the statement current balance.'),
 ('Ocean Harbor','Pearl_Holding_Statement.pdf - Pearl Holding Group, producer code 6883. 6 transactions, ties to Commission Due: $326.21.'),
 ('National General','National_General_Commission_Statement_DB_June_2026.xlsx - Princeton code 9019644. 18 transactions, ties to Grand Total Commission: $871.27.'),
 ('AmWins','Amwins_Comm_1.PDF - Amwins Specialty Auto of Florida, agent 246500. 9 rows, ties to the Agency Total: $72.34.'),
 ('GEICO','Commission_Statement_35.pdf - GEICO, payee Universal Brokers LLC. 38 transactions, ties to the Payment Amount: $2,267.79.'),
 ('',''),
 ('How commission is calculated',''),
 ('Matching','Every binder-book policy was matched to the carrier statements by POLICY NUMBER, not by name, so name spelling differences between the binder and the carriers do not affect the result.'),
 ('Rate used','The rate is taken from the carrier statement itself - it is never assumed. Column K of the Transaction Detail tab is the commission the carrier printed; column L re-performs it as basis x rate and column M shows any difference.'),
 ('Progressive basis','Gross premium on the Progressive detail line. Progressive shows a per-policy rate that varies: 8%, 9%, 10%, 12% and 14% all appear.'),
 ('Infinity basis','Premium column on the Kemper statement. Kemper shows a flat 0.10 (10%) on every FL line.'),
 ('United Auto basis','Transaction amount on the United statement, booked at 10% AS COLLECTED.'),
 ('United Auto 3% incentive','United also pays a separate 3% PROMOTIONAL INCENTIVE line on most policies. By instruction this workbook books United at 10% only, so those incentive lines are EXCLUDED from every Doral total. They are still carried on the Transaction Detail tab exactly as United shows them (column K, shaded), and column O zeroes them so they never reach the Commission Summary. The amount excluded for Doral clients is $252.85 - if you later decide to book it, that is what it would add.'),
 ('United Auto as collected','Most United policies are direct bill and pay COMM AS COLLECTED, meaning June commission is earned only on the premium United actually collected in June - not on the full term premium. That is why, for example, Luis A. Ortiz shows $163.60 commissionable against a $1,103.35 binder premium. The rest of that policy will earn commission on later monthly statements. The four United new-business policies (Gustave, Cordero, Vazquez, Howard) are shown on full premium.'),
 ('United Auto net','United\'s statement nets to $444.33 for the whole agency because cancellations on non-Doral policies charge back more than $1,300. The Doral book on its own earned $902.53 at 10%, which is why its share of that statement reads over 100% on the Carrier Recap tab.'),
 ('United Auto signs','United prints amounts owed TO the agency with a trailing minus and chargebacks as positives. Those signs are reversed in this workbook so that commission earned is positive and chargebacks are negative, which is why the sum here reads $444.33 rather than $444.33-.'),
 ('Ocean Harbor basis','Premium only. The $35.00 policy fee that Pearl lists separately is NOT commissionable - e.g. Javier Forero: $1,153.00 x 11% = $126.83, with the $35.00 fee excluded. The binder premium of $1,188.00 is premium plus fee, which is why it is higher than the commissionable premium.'),
 ('National General basis','Written premium on the Summary Details tab. National General shows a flat 10%.'),
 ('AmWins basis','NET CASH (gross cash less fees), not written premium. AmWins shows 10% and applies it to net cash - e.g. Angely Gejo: $164.67 net cash x 10% = $16.47.'),
 ('Binder vs commissionable','Binder Premium and Commissionable Premium are not meant to agree. The binder figure is the full term premium including MVR and policy fees; the commissionable figure is what the carrier actually paid on this month - net of fees, net of endorsements and cancellations, and for United only the premium collected in June.'),
 ('Multiple transactions','Where a policy has more than one line in the month, all lines are netted and the Notes column on the summary lists them. United incentive lines repeat the same premium as their 10% line, so they are excluded from Commissionable Premium as well (column N of the detail tab) and the premium is never double counted.'),
 ('',''),
 ('Items with no June commission',''),
 ('Not on the June statement','6 Progressive renewals on the binder book do not appear anywhere in the June Progressive detail: Amarillys Gonzalez (970306264), Brudys Garcia (990100610), Yosvany Larralde (982064065), Ana G Castano (866496112), Manuel Martinez (990228513), Reysel Castillo (970570633). Verified by both policy number and name. Most likely the commission fell into the May or July statement period - worth following up with Progressive.'),
 ('No June activity','5 policies are 12-month contracts effective in late 2025 that carry no premium on the binder sheet and show no June 2026 activity on their carrier statement, so $0 is the correct answer for them this month, not a missing match: Levy Diaz Torres, Elizabeth Mirabal Hernandez, Ana Maria Acosta and James John Ciullo (all Infinity), plus Armando Caralos (National General, policy 2032597919).'),
 ('GEICO','No client on the June binder sheet appears on the GEICO statement, but the wider Doral book does: Iris Lopez Sanchez (policy 6219586788) has two renewal-year lines, netting -$45.78. Those two are on the Book Activity by Carrier tab. The remaining 36 GEICO transactions belong to writing agents Alberto Manzor Jr and Amanda Montano and contribute $0 to the Doral totals.'),
 ('',''),
 ('Book Activity by Carrier tab',''),
 ('What it is','Every transaction on the June carrier statements that belongs to a policy in the wider Doral book of business but is NOT on the June binder sheet: endorsements, cancellations, credit endorsements, uncollected premium and as-collected commission on policies written in earlier months. Grouped into a section per carrier, each with its own subtotal.'),
 ('How it was matched','All 384 statement transactions were tested against all 1,383 distinct policy numbers in the full book. Matching is by policy number, allowing for the suffixes carriers append (National General adds a two-digit term, GEICO appends a second number after a dash, United pads with leading zeros). Every match was then cross-checked on surname between the statement insured and the book customer - all 171 matches agreed, and none had to be resolved by judgement.'),
 ('What it shows','103 transactions across all seven carriers, netting -$1,860.29 as the carriers printed it, or -$1,593.97 once United\'s 3% incentive lines are excluded under the 10% rule. It is negative because most of this activity is cancellations and chargebacks on business written in earlier months - real money already collected and now being clawed back, which the June binder sheet alone does not show.'),
 ('United treatment','The same 10%-as-collected rule applies here. Incentive lines are listed and shaded but zeroed in the Commission Counted column.'),
 ('Not Doral business','213 of the 384 statement transactions matched no policy in the Doral book. Those belong to the agency\'s other offices and are excluded from every Doral figure in this workbook.'),
 ('Relationship to the summary','The Commission Summary now carries this activity as its own section below the June binder rows, so the sheet foots to the real number. The June binder sheet total of $6,361.68 is still shown on its own line above it.'),
 ('',''),
 ('Statement adjustments and fees',''),
 ('What they are','Chargebacks and fees the carriers apply OUTSIDE the policy transaction listing, on separate sections of their statements. Three carriers have them this month: National General -$62.10 of loss-history and violation-history chargebacks, Kemper -$32.90 of UW report fees, and Pearl -$16.72 of MVR costs. Total -$111.72, of which -$43.42 is attributable to the Doral book.'),
 ('Why they are matched differently','These sections carry no policy number - National General gives a quote number, Pearl abbreviates the name to a surname and initial, and Kemper names no insured at all. They are therefore matched by NAME, which is weaker than the policy-number matching used everywhere else in this workbook. Each row on the Book Activity tab shows how it was matched, and only the rows marked YES are counted.'),
 ('National General','8 quote numbers matched the Doral book, worth -$35.06. Violation-history rows share a quote number with a loss-history row, so where the quote is Doral every row on it is treated as Doral - that is how the two drivers on quote 242804983 are picked up.'),
 ('Kemper','The -$32.90 is an agency-level UW report fee with no insured named, so none of it can be attributed to the Doral book. It is listed for completeness and counted as $0.'),
 ('Pearl','JOHNSON, J is matched to Jahmar J Johnson, who appears on the same statement as a Doral cancellation, worth -$8.36. MARTINEZ, L could not be identified and is not counted.'),
 ('',''),
 ('Controls performed',''),
 ('Binder premium tie-out','The Binder Premium column reproduces the binder sheet\'s own totals exactly: new business $29,109.40 ($28,891.00 base + $218.40 MVR/fees) and renewals $49,272.55, confirming all 63 rows were captured.'),
 ('Carrier tie-out','Each statement was re-added from its individual transaction lines and agreed to the total the carrier printed: Progressive $14,289.02, Kemper/Infinity $656.60, United Auto $444.33, Ocean Harbor $326.21, National General $871.27, AmWins $72.34, GEICO $2,267.79.'),
 ('Cross-check on Ocean Harbor','Pearl Holding\'s statement is a scanned image, so it was read by OCR. The reading is confirmed three ways: the six lines re-add to the printed Commission Due of $326.21, every line reproduces as premium x rate, and both Doral policy numbers and premiums agree with the binder sheet ($1,153 + $35 fee = the $1,188 on the binder; $1,141 + $35 = $1,176).'),
 ('Rate re-performance','Column M of the Transaction Detail tab is the difference between the carrier-printed commission and basis x rate. It is $0.00 on every one of the 384 transactions except seven United Auto promotional-incentive lines, where United truncates the 3% to the cent instead of rounding it. Because the incentive is excluded from the Doral totals, every transaction that does reach those totals re-performs exactly, to the cent.'),
 ('Agency codes','The six statements are issued to different agency codes and letterheads (Universal Brokers LLC, Creative Insurance Agency, producer code 6883, Princeton code 9019644). Every Doral policy number nevertheless matched its carrier statement exactly, so the codes are alternate identities for the same book rather than a mismatch.'),
 ('',''),
 ('Reading the workbook',''),
 ('Commission Summary','Four blocks. New Business and Renewals are the June binder sheet, footing to $6,361.68. Below them, Other Doral Book Activity carries one row per wider-book policy with June statement activity (-$1,593.97), and Statement Adjustments & Fees carries one row per carrier (-$43.42). The bottom line, Total Doral June Commission, is $4,724.29.'),
 ('Transaction Detail','Every line from all seven statements. Column P flags whether the line belongs to a Doral binder-book client. Blue figures are taken straight from the carrier statement; black figures are calculated. Columns N and O are what actually feeds the Commission Summary - United incentive lines sit at zero there.'),
 ('Carrier Recap','Per-carrier totals, what share of each carrier statement the June binder book represents, and the other book activity added alongside it to give a total Doral commission per carrier.'),
 ('Book Activity by Carrier','A section per carrier listing the wider-book transactions described above, each with a subtotal, and a grand total at the foot.'),
]
nt.cell(1,1).value='Notes, Sources & Methodology'
rn=3
for a,b_ in notes:
    nt.cell(rn,1).value=a; nt.cell(rn,2).value=b_
    if a and not b_: nt.cell(rn,1).font=Font(name=FONT,bold=True,size=11,color=NAVY)
    rn+=1

# ================= Formatting =================
def base_style(ws):
    for row in ws.iter_rows():
        for c in row:
            if not c.font.bold: c.font=Font(name=FONT,size=10)
            c.alignment=Alignment(vertical='center')

for i,t in enumerate(title,1):
    s.cell(i,1).font=Font(name=FONT,size=14 if i==1 else 10,bold=(i==1),color=NAVY if i==1 else '404040')
base_style(s)
for i,t in enumerate(title,1):
    s.cell(i,1).font=Font(name=FONT,size=14 if i==1 else 10,bold=(i==1),
                          italic=(i==3),color=NAVY if i==1 else '404040')
for j,w in enumerate([10,34,12,14,17,17,18,11,9,14,38,7,15,16,16,11,95],1):
    s.column_dimensions[get_column_letter(j)].width=w
for j in range(1,18):
    c=s.cell(HDR_ROW,j); c.fill=HDR; c.font=Font(name=FONT,size=10,bold=True,color='FFFFFF')
    c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center'); c.border=box
s.freeze_panes=s.cell(HDR_ROW+1,3); s.row_dimensions[HDR_ROW].height=30
for i in range(HDR_ROW+1,SUM_LAST+1):
    for j in (10,14,15): s.cell(i,j).number_format=MONEY
    s.cell(i,16).number_format=PCT
    for j in (9,12,13): s.cell(i,j).alignment=Alignment(horizontal='center')
    if isinstance(s.cell(i,13).value,float): s.cell(i,13).number_format=PCT
    if (s.cell(i,17).value or '').startswith('Policy not found'):
        for j in range(1,18): s.cell(i,j).fill=WARN
s.column_dimensions['R'].hidden=True; s.cell(HDR_ROW,18).value='key'
s.auto_filter.ref=f'A{HDR_ROW}:Q{JUNE_LAST}'

base_style(d)
for j,w in enumerate([26,17,40,20,32,22,14,17,17,12,17,15,9,15,16,12,30],1):
    d.column_dimensions[get_column_letter(j)].width=w
for j in range(1,18):
    c=d.cell(1,j); c.fill=HDR; c.font=Font(name=FONT,size=10,bold=True,color='FFFFFF')
    c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center'); c.border=box
d.freeze_panes=d.cell(2,3); d.row_dimensions[1].height=42
d.auto_filter.ref=f'A1:Q{n+1}'
for i in range(2,n+2):
    for j in (8,9,11,12,13,14,15): d.cell(i,j).number_format=MONEY
    d.cell(i,10).number_format=PCT
    d.cell(i,16).alignment=Alignment(horizontal='center')
    for j in (8,9,10,11): d.cell(i,j).font=BLUE
    if d.cell(i,16).value=='YES':
        for j in range(1,18): d.cell(i,j).fill=GREEN
    if d.cell(i,6).value==PROMO:
        for j in (6,15): d.cell(i,j).fill=WARN
for rw in (tr,tr+1):
    for j in range(1,18):
        d.cell(rw,j).fill=TOT; d.cell(rw,j).font=Font(name=FONT,size=10,bold=True)
        if j in (8,11,12,13,14,15): d.cell(rw,j).number_format=MONEY
d.cell(1,11).comment=Comment('Blue figures are taken directly from the carrier commission statement. '
  'Column L re-performs them as basis x rate and column M shows the difference.','Analysis')
d.cell(1,15).comment=Comment('United Auto is booked at 10% as collected. Its separate 3% promotional '
  'incentive lines (shaded) are carried in column K exactly as United shows them, but are zeroed here so '
  'they do not reach the Commission Summary.','Analysis')

base_style(rc); rc.cell(1,1).font=Font(name=FONT,size=13,bold=True,color=NAVY)
for j,w in enumerate([20,40,32,12,14,20,20,24,14,15,20,20,20],1): rc.column_dimensions[get_column_letter(j)].width=w
for j in range(1,14):
    c=rc.cell(3,j); c.fill=HDR; c.font=Font(name=FONT,size=10,bold=True,color='FFFFFF')
    c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center'); c.border=box
rc.row_dimensions[3].height=42
for i in range(4,grr+1):
    for j in (6,7,8,11,12,13): rc.cell(i,j).number_format=MONEY
    rc.cell(i,9).number_format=PCT
    for j in (4,5,10): rc.cell(i,j).alignment=Alignment(horizontal='center')

base_style(ba)
ba.cell(1,1).font=Font(name=FONT,size=13,bold=True,color=NAVY)
ba.cell(2,1).font=Font(name=FONT,size=10,italic=True,color='404040')
ba.cell(2,1).alignment=Alignment(vertical='center',wrap_text=True)
ba.merge_cells(start_row=2,start_column=1,end_row=2,end_column=11); ba.row_dimensions[2].height=30
for j,w in enumerate([32,20,28,22,14,15,12,17,16,13,18],1): ba.column_dimensions[get_column_letter(j)].width=w
for j in range(1,12):
    c=ba.cell(BA_HDR,j); c.fill=HDR; c.font=Font(name=FONT,size=10,bold=True,color='FFFFFF')
    c.alignment=Alignment(vertical='center',wrap_text=True,horizontal='center'); c.border=box
ba.row_dimensions[BA_HDR].height=32; ba.freeze_panes=ba.cell(BA_HDR+1,3)
for i in range(BA_HDR+1,BA_TOT+1):
    for j in (6,8,9): ba.cell(i,j).number_format=MONEY
    ba.cell(i,7).number_format=PCT
    for j in (6,7,8): 
        if isinstance(ba.cell(i,j).value,(int,float)): ba.cell(i,j).font=BLUE
    if ba.cell(i,4).value==PROMO:
        for j in (4,9): ba.cell(i,j).fill=WARN
ba.auto_filter.ref=f'A{BA_HDR}:K{BA_TOT-1}'

nt.cell(1,1).font=Font(name=FONT,size=13,bold=True,color=NAVY)
nt.column_dimensions['A'].width=26; nt.column_dimensions['B'].width=140
for row in nt.iter_rows():
    for c in row:
        if not c.font.bold: c.font=Font(name=FONT,size=10)
        c.alignment=Alignment(vertical='top',wrap_text=True)

wb._sheets=[wb[t] for t in ['Commission Summary','Transaction Detail','Carrier Recap',
                            'Book Activity by Carrier','Notes & Sources']]
for ws in wb: ws.sheet_view.showGridLines=False
wb.save('Doral_Binder_Book_June_2026_Commissions.xlsx')
print("saved; detail rows:",n)
