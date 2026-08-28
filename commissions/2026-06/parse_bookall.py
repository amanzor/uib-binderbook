import openpyxl, json, re, datetime
SRC="/root/.claude/uploads/1c03ef55-29f7-5482-9b94-43f345ea2246/a23d71f3-Doral_Office_Binder_Book.xlsx"
wb=openpyxl.load_workbook(SRC,data_only=True)

def norm(p):
    p=re.sub(r'[^A-Z0-9]','',str(p).upper())
    m=re.match(r'^([A-Z]*)(\d*)$',p)
    if m and m.group(2): return m.group(1)+m.group(2).lstrip('0')
    return p

def s(v):
    if v is None: return ''
    if isinstance(v,datetime.datetime): return v.strftime('%m/%d/%Y')
    return str(v).strip()

recs=[]; sheets_used=[]
for ws in wb:
    hdr=None
    for r in ws.iter_rows(min_row=1,max_row=min(ws.max_row,6),values_only=True):
        if r and len(r)>1 and s(r[1]).lower()=='customer name': hdr=True; break
    if not hdr: continue
    sheets_used.append(ws.title); cnt=0
    for i,r in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
        r=list(r)+['']*20
        name=s(r[1]); pol=s(r[13]); co=s(r[5])
        if not name or name.lower() in ('customer name','','total'): continue
        if not pol: continue
        if re.fullmatch(r'[A-Za-z ,.\'-]+',pol): continue        # notes like "need policy number"
        recs.append(dict(sheet=ws.title,row=i,name=name,status=s(r[2]),ptype=s(r[3]),
                         lob=s(r[4]),co=co,base=s(r[9]),total=s(r[10]),term=s(r[11]),
                         eff=s(r[12]),policy=pol,key=norm(pol)))
        cnt+=1
    print(f"  {ws.title:22} {cnt:5} policies")
print("\nsheets with book layout:",len(sheets_used))
print("total policy rows:",len(recs))
uniq={}
for x in recs:
    uniq.setdefault(x['key'],[]).append(x)
print("distinct policy numbers:",len(uniq))
json.dump(recs,open('book_all.json','w'),indent=1)
