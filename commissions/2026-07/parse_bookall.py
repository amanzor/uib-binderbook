"""Parse every monthly sheet in the binder book into a flat list of policy
records (book_all.json). Used later to match carrier-statement transactions to
the wider Doral book of business by policy number, the same way June did."""
import openpyxl, json, re, datetime

SRC = "Doral_Office_Binder_Book.xlsx"


def norm(p):
    p = re.sub(r'[^A-Z0-9]', '', str(p).upper())
    m = re.match(r'^([A-Z]*)(\d*)$', p)
    if m and m.group(2):
        return m.group(1) + m.group(2).lstrip('0')
    return p


def s(v):
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%m/%d/%Y')
    return str(v).strip()


wb = openpyxl.load_workbook(SRC, data_only=True)
recs = []
sheets_used = []
for ws in wb:
    hdr = False
    for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 6), values_only=True):
        if r and len(r) > 1 and s(r[1]).lower() == 'customer name':
            hdr = True
            break
    if not hdr:
        continue
    sheets_used.append(ws.title)
    cnt = 0
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        r = list(r) + [''] * 20
        name, pol, co = s(r[1]), s(r[13]), s(r[5])
        if not name or name.lower() in ('customer name', '', 'total'):
            continue
        if not pol:
            continue
        if re.fullmatch(r'[A-Za-z ,.\'-]+', pol):        # notes like "need policy number"
            continue
        recs.append(dict(sheet=ws.title, row=i, name=name, status=s(r[2]),
                         ptype=s(r[3]), lob=s(r[4]), co=co, base=s(r[9]),
                         total=s(r[10]), term=s(r[11]), eff=s(r[12]),
                         policy=pol, key=norm(pol)))
        cnt += 1

json.dump(recs, open('book_all.json', 'w'), indent=1)
uniq = {}
for x in recs:
    uniq.setdefault(x['key'], []).append(x)
print(f"sheets with book layout: {len(sheets_used)}")
print(f"total policy rows:       {len(recs)}")
print(f"distinct policy numbers: {len(uniq)}")
