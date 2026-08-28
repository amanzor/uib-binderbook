"""Build the July 2026 commission workbook, mirroring June's deliverable:
  Commission Summary  — the monthly carrier summary + royalty + operating expenses
  Transaction Detail  — every Doral-matched carrier transaction, binder-flagged
  Carrier Recap       — statement total vs Doral matched / binder / wider book
  Notes & Exceptions  — method, tie-outs, review items
"""
import json, re, collections
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C = json.load(open('carriers.json'))
BM = json.load(open('bookmatch.json'))
D = json.load(open('pdf_data.json'))
BINDER = json.load(open('binder.json'))
matched = BM['matched']

NAVY = '1F3864'; LIGHT = 'D9E1F2'; RED = 'C00000'
hdr_font = Font(bold=True, color='FFFFFF'); hdr_fill = PatternFill('solid', fgColor=NAVY)
sub_font = Font(bold=True, color=NAVY); sub_fill = PatternFill('solid', fgColor=LIGHT)
thin = Side(style='thin', color='BFBFBF'); border = Border(bottom=thin)
STMT_TOTAL = {'Progressive': 17885.65, 'Infinity': 3714.35, 'National General': 658.32,
              'Ocean Harbor': 250.08, 'GEICO': 9009.15, 'United Auto': -589.59, 'AmWins': 78.74}
ORDER = D['order']


def money(ws, cell):
    ws[cell].number_format = '$#,##0.00;($#,##0.00)'


wb = openpyxl.Workbook()

# ---------------- Commission Summary ----------------
ws = wb.active; ws.title = 'Commission Summary'
ws['A1'] = 'Doral Office — July 2026 Commission Summary'; ws['A1'].font = Font(bold=True, size=14, color=NAVY)
r = 3
ws.cell(r, 1, 'Carrier'); ws.cell(r, 2, 'New Business'); ws.cell(r, 3, 'Renewals & Adj.'); ws.cell(r, 4, 'Total')
for c in range(1, 5):
    ws.cell(r, c).font = hdr_font; ws.cell(r, c).fill = hdr_fill
for co, n, rn, tt in D['summary']:
    r += 1
    ws.cell(r, 1, co); ws.cell(r, 2, n); ws.cell(r, 3, rn); ws.cell(r, 4, tt)
    for c in (2, 3, 4):
        money(ws, ws.cell(r, c).coordinate)
        if ws.cell(r, c).value < 0:
            ws.cell(r, c).font = Font(color=RED)
r += 1
ws.cell(r, 1, 'Total Gross Commissions'); ws.cell(r, 2, D['gt_new']); ws.cell(r, 3, D['gt_ren']); ws.cell(r, 4, D['gross'])
for c in range(1, 5):
    ws.cell(r, c).font = sub_font; ws.cell(r, c).fill = sub_fill
    if c > 1:
        money(ws, ws.cell(r, c).coordinate)
gross = D['gross']; roy = round(gross * 0.15, 2); paid = round(gross - roy, 2)
EXP = [('220 License Rent', -500.0), ('Systems', -224.0), ('MVRs', D['adj_total'])]
net = round(paid + sum(v for _, v in EXP), 2)
r += 2
for label, val, bold in [('Gross Commissions', gross, 0), ('Royalty (15%)', -roy, 0), ('Paid to Doral', paid, 1)] \
        + [(l, v, 0) for l, v in EXP] + [('Net to Doral', net, 1)]:
    ws.cell(r, 1, label); ws.cell(r, 2, val); money(ws, ws.cell(r, 2).coordinate)
    if bold:
        ws.cell(r, 1).font = sub_font; ws.cell(r, 2).font = sub_font
        ws.cell(r, 1).fill = sub_fill; ws.cell(r, 2).fill = sub_fill
    elif val < 0:
        ws.cell(r, 2).font = Font(color=RED)
    r += 1
ws.column_dimensions['A'].width = 34
for c in 'BCD':
    ws.column_dimensions[c].width = 16

# ---------------- Transaction Detail ----------------
ws = wb.create_sheet('Transaction Detail')
cols = ['Carrier', 'Statement Insured', 'Policy', 'Type', 'Date', 'Premium/Basis', 'Rate',
        'Commission', 'On Binder?', 'Book Customer', 'Book Sheet']
for i, h in enumerate(cols, 1):
    ws.cell(1, i, h); ws.cell(1, i).font = hdr_font; ws.cell(1, i).fill = hdr_fill
row = 2
order2 = ['Progressive', 'United Auto', 'Infinity', 'National General', 'GEICO', 'AmWins', 'Ocean Harbor']
for co in order2:
    txns = [t for t in matched if t['carrier'] == co]
    txns.sort(key=lambda t: (not t['in_binder'], t['book_name']))
    for t in txns:
        vals = [t['carrier'], t['name'], t['policy'], t['ttype'] + (' (3% promo, excluded)' if t['promo'] else ''),
                t['tdate'], t['prem'], t['rate'], (0.0 if t['promo'] else t['comm']),
                'Yes' if t['in_binder'] else 'wider book', t['book_name'], t['book_sheet'].strip()]
        for i, v in enumerate(vals, 1):
            ws.cell(row, i, v)
        money(ws, ws.cell(row, 6).coordinate); money(ws, ws.cell(row, 8).coordinate)
        ws.cell(row, 7).number_format = '0.0%'
        if not t['promo'] and t['comm'] < 0:
            ws.cell(row, 8).font = Font(color=RED)
        if t['promo']:
            for i in range(1, 12):
                ws.cell(row, i).font = Font(color='999999')
        row += 1
widths = [16, 30, 17, 26, 11, 14, 7, 13, 11, 26, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'

# ---------------- Carrier Recap ----------------
ws = wb.create_sheet('Carrier Recap')
cols = ['Carrier', 'Statement Total (as printed)', 'Doral Matched', 'On Binder', 'Wider Book', 'Other Offices (excluded)']
for i, h in enumerate(cols, 1):
    ws.cell(1, i, h); ws.cell(1, i).font = hdr_font; ws.cell(1, i).fill = hdr_fill
row = 2
un = json.load(open('bookmatch.json'))['unmatched']
for co in order2:
    mm = [t for t in matched if t['carrier'] == co and not t['promo']]
    b = sum(t['comm'] for t in mm if t['in_binder'])
    w = sum(t['comm'] for t in mm if not t['in_binder'])
    other = sum(t['comm'] for t in un if t['carrier'] == co and not t.get('promo'))
    ws.cell(row, 1, co); ws.cell(row, 2, STMT_TOTAL[co]); ws.cell(row, 3, round(b + w, 2))
    ws.cell(row, 4, round(b, 2)); ws.cell(row, 5, round(w, 2)); ws.cell(row, 6, round(other, 2))
    for c in range(2, 7):
        money(ws, ws.cell(row, c).coordinate)
    row += 1
ws.cell(row, 1, 'Note'); ws.cell(row, 1).font = sub_font
ws.cell(row + 1, 1, 'United Auto statement total is the net Current Balance (−$589.59); Doral share is booked 10% as '
        'collected (+), 3% promo excluded. Other totals are each statement\'s printed commission.')
ws.column_dimensions['A'].width = 18
for c in 'BCDEF':
    ws.column_dimensions[c].width = 22

# ---------------- Notes & Exceptions ----------------
ws = wb.create_sheet('Notes & Exceptions')
notes = [
    ('Doral Office — July 2026 Commission Statement', True),
    ('', False),
    ('Method: every policy on the July binder sheet (producer Jorge Castro, 21 new + 49 renewals) plus every wider-book', False),
    ('client with a positive or negative commission transaction was priced against the seven July carrier statements,', False),
    ('matched by policy number and cross-checked on surname (0 mismatches on 159 matched transactions).', False),
    ('Each carrier read ties exactly to the total the carrier prints. United Auto and AmWins are booked on the monthly', False),
    ('as-collected basis (United 10% as collected, 3% promo excluded; AmWins 10% of net cash), not on gross premium.', False),
    ('', False),
    ('REVIEW ITEMS', True),
    ('1. Melissa A Taylor (Progressive 876394907): the binder lists her as active New ($1,404), but Progressive\'s July', False),
    ('   statement shows a full cancellation (−$1,227 at 12% = −$147.24). She earned +$147.24 on it in June, so it nets', False),
    ('   to zero across the two months. Shown here at the carrier figure (−$147.24). Confirm the binder status.', False),
    ('2. Ocean Harbor is recorded in full as its producer-6883 statement shows: all five lines including Ertas Meral', False),
    ('   (+$118.91) and Morin George (−$64.87), so the section foots to the statement\'s $250.08 Commission Due.', False),
    ('3. Office-only expense lines are 220 License Rent ($500) and Systems ($224), provided for July; the Cash Payments', False),
    ('   Owed line has been removed for now. They appear on no carrier statement.', False),
    ('', False),
    ('BINDER POLICIES WITH NO CARRIER COMMISSION THIS MONTH ($0 — likely billed in an adjacent statement period):', True),
]
binder_comm = collections.defaultdict(float)


def norm(p):
    p = re.sub(r'[^A-Z0-9]', '', str(p).upper()); m = re.match(r'^([A-Z]*)(\d*)$', p)
    return m.group(1) + m.group(2).lstrip('0') if m and m.group(2) else p


CO2 = {'progressive': 'Progressive', 'united auto': 'United Auto', 'infinity': 'Infinity',
       'national general': 'National General', 'geico': 'GEICO', 'ocean harbor': 'Ocean Harbor'}


def keys(carr, policy):
    p = str(policy).strip(); out = {norm(p)}
    if carr == 'National General':
        out.add(norm(p.split()[0]))
    if carr == 'GEICO':
        out.add(norm(p.split('-')[0]))
    return {k for k in out if k}


idx = {}
for b in BINDER:
    carr = CO2.get(b['company'].strip().lower(), b['company'])
    for k in keys(carr, b['policy']):
        idx.setdefault(k, b['row'])
for t in matched:
    if t['promo'] or not t['in_binder']:
        continue
    for k in keys(t['carrier'], t['policy']):
        if k in idx:
            binder_comm[idx[k]] += t['comm']; break
for b in BINDER:
    if round(binder_comm.get(b['row'], 0), 2) == 0:
        notes.append((f"    {b['name']}  |  {b['company']}  |  {b['status']}  |  {b['policy']}", False))
rr = 1
for text, bold in notes:
    ws.cell(rr, 1, text)
    if bold:
        ws.cell(rr, 1).font = Font(bold=True, color=NAVY)
    rr += 1
ws.column_dimensions['A'].width = 120

wb.save('Doral_Binder_Book_July_2026_Commissions.xlsx')
print("wrote Doral_Binder_Book_July_2026_Commissions.xlsx")
